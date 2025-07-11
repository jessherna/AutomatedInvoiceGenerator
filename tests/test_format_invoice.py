import pytest
from invoice import FormatInvoice
from openpyxl import Workbook
from datetime import datetime, timedelta

@pytest.fixture
def sample_order():
    return {
        "InvoiceNumber": "INV-2024-001",
        "InvoiceDate": datetime.now(),
        "DueDate": datetime.now() + timedelta(days=30),
        "ItemID": "ABC123",
        "Qty": 2,
        "Price": 9.99,
        "CustomerName": "Acme Corp",
        "Email": "acme@example.com",
        "Items": [
            {
                "Qty": 2,
                "Description": "Test Item",
                "UnitPrice": 9.99
            }
        ]
    }

def test_format_invoice_returns_workbook(sample_order):
    wb = FormatInvoice(sample_order)
    assert isinstance(wb, Workbook)
    assert "Invoice" in wb.sheetnames

def test_format_invoice_header_and_items(sample_order):
    wb = FormatInvoice(sample_order)
    ws = wb["Invoice"]
    
    # Check title
    assert ws["B4"].value == "INVOICE"
    
    # Check company info
    assert ws["A9"].value == "Yukon Packing"
    
    # Check metadata
    assert ws["E14"].value == "Invoice #"
    assert ws["F14"].value == sample_order["InvoiceNumber"]
    
    # Check line items table headers
    headers = [ws.cell(row=21, column=i).value for i in range(1,7)]
    assert headers == ["Qty", "Description", "Unit Price", "Amount", "Notes", "Status"]
    
    # Check line item data
    item = sample_order["Items"][0]
    assert ws.cell(row=22, column=1).value == item["Qty"]
    assert ws.cell(row=22, column=2).value == item["Description"]
    assert ws.cell(row=22, column=3).value == item["UnitPrice"]
    
    # Check total calculation
    expected_total = item["Qty"] * item["UnitPrice"]
    expected_with_gst = round(expected_total * 1.05, 2)  # Including 5% GST, rounded to 2 decimal places
    
    # Find the TOTAL row (it should be the last row of the summary section)
    total_row = None
    for row in range(21, 30):  # Look in a reasonable range
        if ws.cell(row=row, column=5).value == "TOTAL":
            total_row = row
            break
    
    assert total_row is not None, "Could not find TOTAL row in invoice"
    assert ws.cell(row=total_row, column=6).value == expected_with_gst

def test_get_next_invoice_number(tmp_path):
    """Test that invoice numbers are sequential and reset daily"""
    from invoice import get_next_invoice_number
    import os
    from datetime import datetime
    
    # Change to temp directory for testing
    original_dir = os.getcwd()
    os.chdir(tmp_path)
    
    try:
        # First call should start with 1
        inv1 = get_next_invoice_number()
        assert inv1.startswith("INV-")
        assert inv1.endswith("-0001")
        
        # Second call should increment
        inv2 = get_next_invoice_number()
        assert inv2.endswith("-0002")
        
        # Third call should increment again
        inv3 = get_next_invoice_number()
        assert inv3.endswith("-0003")
        
        # Verify the counter file exists and has correct format
        with open("invoice_counter.txt", "r") as f:
            date, counter = f.read().strip().split("-")
            assert date == datetime.now().strftime("%Y%m%d")
            assert int(counter) == 3
            
    finally:
        # Clean up and return to original directory
        os.chdir(original_dir)
