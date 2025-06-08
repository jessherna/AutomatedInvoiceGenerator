from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Font, Border, Side
import os
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import platform


def LoadOrders(path: str) -> list[dict]:
    """
    Reads the Orders sheet and returns a list of dicts. 
    Expects columns: CustomerID, ItemID, Qty, Price, CustomerName, Email
    """
    wb = load_workbook(path, data_only=True)
    ws = wb["Orders"]
    rows = list(ws.values)
    headers = rows[0]
    orders = []
    for row in rows[1:]:
        entry = dict(zip(headers, row))
        # convert numeric fields
        if "Qty" in entry:
            entry["Qty"] = int(entry["Qty"])
        if "Price" in entry:
            entry["Price"] = float(entry["Price"])
        orders.append(entry)
    return orders

def LoadBillToData(path: str) -> dict:
    """
    Reads the BillTo sheet and returns a dict mapping CustomerID to its row data.
    Expects columns: CustomerID, CustomerName, Email, Phone, Address, City
    
    If a customer is not found in the bill-to data, returns an empty dict for that customer.
    """
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["BillTo"]
        rows = list(ws.values)
        headers = rows[0]
        data = {}
        for row in rows[1:]:
            entry = dict(zip(headers, row))
            if "CustomerID" in entry and entry["CustomerID"]:  # Only add if CustomerID exists and is not empty
                data[entry["CustomerID"]] = entry
        return data
    except (KeyError, FileNotFoundError):
        print(f"Warning: Could not load bill-to data from {path}. Using empty bill-to data.")
        return {}

def LoadItemData(path: str) -> dict:
    """
    Reads the Items sheet and returns a dict mapping ItemID to item details.
    Expects columns: ItemID, Name, Description, UnitPrice
    """
    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["Items"]
        rows = list(ws.values)
        headers = rows[0]
        data = {}
        for row in rows[1:]:
            entry = dict(zip(headers, row))
            if "ItemID" in entry and entry["ItemID"]:  # Only add if ItemID exists and is not empty
                data[entry["ItemID"]] = entry
        return data
    except (KeyError, FileNotFoundError):
        print(f"Warning: Could not load item data from {path}. Using empty item data.")
        return {}

def FormatInvoice(order: dict) -> Workbook:
    """
    Generate a professional, black-and-white invoice with a styled table and currency formatting.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # — Logo & Title —
    try:
        logo = Image("docs/assets/logo.png")
        logo.width, logo.height = 150, 150
        ws.add_image(logo, "A1")
    except FileNotFoundError:
        pass

    ws.merge_cells("B4:F4")
    title = ws["B4"]
    title.value = "INVOICE"
    title.font = Font(size=20, bold=True, color="000000")
    title.alignment = Alignment(horizontal="right", vertical="top")

    # — Company Info & Contact —
    company_lines = [
        "Yukon Packing",
        "443 Maple Avenue",
        "Ontario, NT B4M 3B7",
        f"Phone: {order.get('CompanyContact','')}"
    ]
    for i, txt in enumerate(company_lines, start=9):
        cell = ws[f"A{i}"]
        cell.value = txt
        cell.font = Font(size=11, color="000000")
        cell.alignment = Alignment(horizontal="left")

    # — Bill To / Ship To —
    ws["A14"].value = "Bill To"
    ws["C14"].value = "Ship To"
    for cell in ("A14","C14"):
        ws[cell].font = Font(size=12, bold=True, color="000000")

    for col, key, col_offset in [("A","BillTo",0), ("C","ShipTo",0)]:
        info = order.get(key, {})
        ws[f"{col}15"].value = info.get("CustomerName","")
        ws[f"{col}16"].value = info.get("Address","")
        ws[f"{col}17"].value = info.get("City","")
        ws[f"{col}18"].value = f"Phone: {info.get('Phone','')}"
        ws[f"{col}19"].value = f"Email: {info.get('Email','')}"
        for r in range(15,20):
            ws[f"{col}{r}"].font = Font(size=11, color="000000")

    # — Metadata —
    meta = {
        "Invoice #":     order["InvoiceNumber"],
        "Invoice Date":  (order["InvoiceDate"].strftime("%d/%m/%Y")
                          if isinstance(order["InvoiceDate"], datetime)
                          else order["InvoiceDate"]),
        "P.O.#":         order.get("PO",""),
        "Due Date":      (order["DueDate"].strftime("%d/%m/%Y")
                          if isinstance(order["DueDate"], datetime)
                          else order["DueDate"])
    }
    row = 14
    for label, val in meta.items():
        lbl_cell = ws[f"E{row}"]
        val_cell = ws[f"F{row}"]
        lbl_cell.value = label
        val_cell.value = val
        lbl_cell.font = Font(bold=True, color="000000")
        val_cell.font = Font(color="000000")
        val_cell.alignment = Alignment(horizontal="right")
        row += 1

    # — Line-Items Table —
    start = 21  # leave a blank row above for breathing room
    headers = ["Qty", "Description", "Unit Price", "Amount", "Notes", "Status"]
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Header row
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=i, value=h)
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Data rows
    for idx, item in enumerate(order["Items"], start=1):
        r = start + idx
        # Qty
        ws.cell(r, 1, item["Qty"]).border = border
        # Description
        ws.cell(r, 2, item["Description"]).border = border
        # Unit Price
        up = ws.cell(r, 3, item["UnitPrice"])
        up.number_format = '"$"#,##0.00'
        up.alignment = Alignment(horizontal="right")
        up.border = border
        # Amount
        amt = item["Qty"] * item["UnitPrice"]
        a = ws.cell(r, 4, amt)
        a.number_format = '"$"#,##0.00'
        a.alignment = Alignment(horizontal="right")
        a.border = border
        # Notes (empty)
        ws.cell(r, 5, "").border = border
        # Status (empty)
        ws.cell(r, 6, "").border = border

    # Turn into a styled Table (black & white)
    last_row = start + len(order["Items"])
    tbl = Table(displayName="InvoiceTable", ref=f"A{start}:F{last_row}")
    tbl_style = TableStyleInfo(
        name="None",  # no styling
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    tbl.tableStyleInfo = tbl_style
    ws.add_table(tbl)

    # — Summary —
    sub = sum(i["Qty"] * i["UnitPrice"] for i in order["Items"])
    gst = round(sub * 0.05, 2)
    summary = [("Subtotal", sub), ("GST 5%", gst), ("TOTAL", sub + gst)]
    base = last_row + 2
    for i, (lbl, val) in enumerate(summary):
        lbl_cell = ws.cell(base + i, 5, lbl)  
        val_cell = ws.cell(base + i, 6, val)  
        lbl_cell.font = Font(bold=True, color="000000")
        lbl_cell.alignment = Alignment(horizontal="right")
        val_cell.number_format = '"$"#,##0.00'
        val_cell.font = Font(color="000000")
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = border

    # — Terms & Conditions —
    trow = base + 4
    ws[f"A{trow}"].value = "Terms & Conditions:"
    ws[f"A{trow}"].font = Font(bold=True, color="000000")
    ws[f"A{trow+1}"].value = order.get("Terms", "Payment is due within 15 days.")
    ws[f"A{trow+1}"].font = Font(color="000000")

    # — Column Widths & Page Setup —
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 12

    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.75, bottom=0.75)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return wb

def ExportInvoice(workbook: Workbook, output_path: str, format: str = "xlsx") -> str:
    """
    Save `workbook` to disk:
      - as .xlsx if format=="xlsx"
      - as .pdf using Excel COM if format=="pdf" (Windows only)
      - as .pdf using LibreOffice if format=="pdf" (Linux)
    Returns the full path of the generated file.
    """
    if format == "xlsx":
        file = output_path + ".xlsx"
        workbook.save(file)
        return file

    elif format == "pdf":
        # First save as xlsx
        xlsx_path = output_path + "_tmp.xlsx"
        workbook.save(xlsx_path)
        
        try:
            # Try Windows method first
            from win32com.client import Dispatch
            excel = Dispatch('Excel.Application')
            excel.Visible = False
            wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
            pdf_path = output_path + ".pdf"
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_path))
            wb.Close()
            excel.Quit()
            os.remove(xlsx_path)
            return pdf_path
        except ImportError:
            # Fall back to LibreOffice on Linux
            try:
                import subprocess
                pdf_path = output_path + ".pdf"
                subprocess.run(['soffice', '--headless', '--convert-to', 'pdf', xlsx_path, '--outdir', os.path.dirname(pdf_path)], check=True)
                os.remove(xlsx_path)
                return pdf_path
            except Exception as e:
                raise RuntimeError(f"PDF conversion failed: {str(e)}")
    else:
        raise ValueError(f"Unsupported format: {format}")
    
def SendInvoice(emailAddr: str, filePath: str, cc: str = None, additional_attachments: list = None) -> None:
    """
    Open Outlook, create a mail item, attach the file at filePath, and send to emailAddr.
    
    Args:
        emailAddr: Email address of the recipient
        filePath: Path to the invoice file to attach
        cc: Optional CC recipient email address
        additional_attachments: Optional list of additional file paths to attach
    """
    if platform.system() != "Windows":
        raise RuntimeError("Email sending is only supported on Windows")

    try:
        from win32com.client import Dispatch
    except ImportError:
        raise RuntimeError("SendInvoice requires pywin32 and Windows COM")

    # Validate and normalize file paths
    def validate_path(path):
        abs_path = os.path.abspath(path)
        if not os.path.exists(abs_path):
            raise FileNotFoundError(f"File not found: {abs_path}")
        return abs_path

    try:
        # Validate main invoice file
        abs_file_path = validate_path(filePath)
        
        # Ensure we're only attaching PDF files
        if not abs_file_path.lower().endswith('.pdf'):
            raise ValueError("Only PDF files can be attached to the email")
        
        outlook = Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)
        mail.To = emailAddr
        if cc:
            mail.CC = cc

        # Extract invoice number from filename
        invoice_number = os.path.basename(filePath).split('.')[0]
        
        mail.Subject = f"Invoice {invoice_number} - Contoso Logistics"
        
        # Create a professional email body
        email_body = f"""Dear Valued Customer,

I hope this email finds you well. Please find attached Invoice {invoice_number} for your recent order with Contoso Logistics.

Invoice Details:
- Invoice Number: {invoice_number}
- Date: {datetime.now().strftime('%B %d, %Y')}

Payment Terms:
- Payment is due within 15 days from invoice date
- Payment methods accepted: Bank Transfer, Credit Card
- For wire transfer details, please refer to the invoice

If you have any questions regarding this invoice or need any clarification, please don't hesitate to contact our accounting department at accounting@contosologistics.com or call us at (416) 555-0123.

Thank you for your business.

Best regards,
Contoso Logistics
Accounting Department
Phone: (416) 555-0123
Email: accounting@contosologistics.com

---
This is an automated message. Please do not reply directly to this email.
For immediate assistance, please contact our customer service department."""

        mail.Body = email_body
        mail.HTMLBody = email_body.replace('\n', '<br>')  # Convert to HTML format

        # Attach main invoice file
        try:
            mail.Attachments.Add(abs_file_path)
        except AttributeError:
            # e.g. DummyMailItem in tests uses Attachments_Add
            if hasattr(mail, "Attachments_Add"):
                mail.Attachments_Add(abs_file_path)
            else:
                raise

        # Attach any additional files (only PDFs)
        if additional_attachments:
            for attachment in additional_attachments:
                abs_attachment_path = validate_path(attachment)
                if not abs_attachment_path.lower().endswith('.pdf'):
                    continue  # Skip non-PDF files
                try:
                    mail.Attachments.Add(abs_attachment_path)
                except AttributeError:
                    if hasattr(mail, "Attachments_Add"):
                        mail.Attachments_Add(abs_attachment_path)
                    else:
                        raise

        mail.Send()
    except FileNotFoundError as e:
        raise RuntimeError(f"Failed to send invoice: File not found - {str(e)}")
    except ValueError as e:
        # Preserve ValueError exceptions
        raise
    except Exception as e:
        raise RuntimeError(f"Failed to send invoice: {str(e)}")

def GenerateAllInvoices(path_to_orders: str) -> None:
    """
    Full pipeline: load all orders, format each invoice, export to PDF,
    and send via Outlook.
    """
    orders = LoadOrders(path_to_orders)
    bill_to_data = LoadBillToData(path_to_orders)
    
    for order in orders:
        # Transform order data
        transformed_order = TransformOrder(order, bill_to_data)
        
        # 1) Format
        wb = FormatInvoice(transformed_order)

        # 2) Export (use OrderID as base filename)
        base = f"invoice_{order.get('OrderID', '')}"
        pdf_path = ExportInvoice(wb, base, format="pdf")

        # 3) Send
        SendInvoice(order.get("Email", ""), pdf_path)

def get_next_po_number() -> str:
    """
    Generates a sequential PO number in the format PO-YYYYMMDD-XXXX
    where XXXX is a sequential number that resets daily.
    
    Returns:
        str: A formatted PO number
    """
    today = datetime.now().strftime("%Y%m%d")
    # In a real application, you would store and retrieve this from a database
    # For now, we'll use a simple file-based approach
    counter_file = "po_counter.txt"
    
    try:
        with open(counter_file, "r") as f:
            last_date, counter = f.read().strip().split("-")
            counter = int(counter)
            if last_date != today:
                counter = 1
            else:
                counter += 1
    except (FileNotFoundError, ValueError):
        counter = 1
    
    with open(counter_file, "w") as f:
        f.write(f"{today}-{counter}")
    
    return f"PO-{today}-{counter:04d}"

def get_next_invoice_number() -> str:
    """
    Generates a sequential invoice number in the format INV-YYYYMMDD-XXXX
    where XXXX is a sequential number that resets daily.
    
    Returns:
        str: A formatted invoice number
    """
    today = datetime.now().strftime("%Y%m%d")
    # In a real application, you would store and retrieve this from a database
    # For now, we'll use a simple file-based approach
    counter_file = "invoice_counter.txt"
    
    try:
        with open(counter_file, "r") as f:
            last_date, counter = f.read().strip().split("-")
            counter = int(counter)
            if last_date != today:
                counter = 1
            else:
                counter += 1
    except (FileNotFoundError, ValueError):
        counter = 1
    
    with open(counter_file, "w") as f:
        f.write(f"{today}-{counter}")
    
    return f"INV-{today}-{counter:04d}"

def TransformOrder(order: dict, bill_to_data: dict) -> dict:
    """
    Transforms a basic order dict into the format expected by FormatInvoice.
    
    Args:
        order: Basic order dict with CustomerID, ItemID, Qty, Price, CustomerName, Email, ItemName
        bill_to_data: Dict mapping CustomerID to bill-to information
        
    Returns:
        Dict in the format expected by FormatInvoice
    """
    # Get bill-to information
    bill_to = bill_to_data.get(order["CustomerID"], {})
    email = bill_to.get("Email", order["Email"])  # Get email from customer's bill-to data
    
    # Create items list
    items = [{
        "Qty": order["Qty"],
        "Description": order.get("ItemName", f"Item {order['ItemID']}"),  # Use ItemName if available
        "UnitPrice": order["Price"]
    }]
    
    # Generate sequential invoice number
    invoice_number = get_next_invoice_number()
    
    # Calculate due date (15 days from now)
    due_date = datetime.now().replace(day=datetime.now().day + 15)
    
    return {
        "CustomerID": order["CustomerID"],
        "CustomerName": order["CustomerName"],
        "Email": email,  # Use bill-to email if available, otherwise use order email
        "BillTo": bill_to,
        "ShipTo": bill_to,  # Using same address for shipping
        "CompanyContact": "555-0123",  # You may want to load this from a config
        "InvoiceNumber": invoice_number,
        "InvoiceDate": datetime.now(),
        "PO": get_next_po_number(),  # Generate sequential PO number
        "DueDate": due_date,
        "Items": items,
        "Terms": "Payment is due within 15 days."
    }
